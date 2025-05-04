#!/usr/bin/env python3
"""
BPL AST to Visio Excel Converter

This script converts a BPL AST JSON file to a Visio-compatible Excel format
for business process visualization.

Usage:
    python ast_to_visio.py input.bpl-ast.json output.xlsx

Dependencies:
    - pandas
    - openpyxl
    - numpy
"""

import json
import sys
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

class BplAstToVisioConverter:
    def __init__(self, ast_json):
        """Initialize with parsed AST json data"""
        self.ast = ast_json
        self.nodes = []
        self.edges = []
        self.swimlanes = []
        
    def normalize_name(self, name):
        """Create a normalized name suitable for Visio"""
        if not name:
            return "unknown"
        return name.lower().replace(" ", "_")
    
    def extract_nodes(self):
        """Extract nodes from AST"""
        for process in self.ast.get('processes', []):
            process_id = process.get('id')
            for lane in process.get('lanes', []):
                lane_id = lane.get('id')
                lane_name = lane.get('name')
                
                # Add lane to swimlanes
                self.swimlanes.append({
                    'id': lane_id,
                    'name': lane_name,
                    'process': process_id
                })
                
                # Process each element in the lane
                for element in lane.get('elements', []):
                    node_type = element.get('type')
                    node_id = element.get('id')
                    node_name = element.get('name')
                    
                    node_data = {
                        'id': node_id,
                        'name': node_name,
                        'type': node_type,
                        'lane': lane_id,
                        'process': process_id
                    }
                    
                    # Add specific attributes based on node type
                    if node_type == 'gateway':
                        node_data['gateway_type'] = element.get('gatewayType', 'exclusive')
                    elif node_type == 'event':
                        node_data['event_type'] = element.get('eventType', 'intermediate')
                    elif node_type == 'task':
                        if 'send:' in node_name:
                            node_data['type'] = 'send'
                        elif 'receive:' in node_name:
                            node_data['type'] = 'receive'
                    
                    self.nodes.append(node_data)
    
    def extract_edges(self):
        """Extract connections from AST"""
        for conn in self.ast.get('connections', []):
            conn_type = conn.get('type')
            source_ref = conn.get('sourceRef')
            target_ref = conn.get('targetRef')
            name = conn.get('name', '')
            
            edge_data = {
                'id': conn.get('id'),
                'type': conn_type,
                'name': name,
                'source': source_ref,
                'target': target_ref
            }
            
            self.edges.append(edge_data)
    
    def process_ast(self):
        """Process the AST and extract all required data"""
        self.extract_nodes()
        self.extract_edges()
        
    def create_visio_dataframes(self):
        """
        Create pandas DataFrames for Visio format
        
        Returns:
            tuple: (nodes_df, edges_df, swimlanes_df)
        """
        # Create DataFrame for nodes
        if self.nodes:
            nodes_df = pd.DataFrame(self.nodes)
            nodes_df['visio_type'] = nodes_df['type'].apply(self.map_node_type_to_visio)
            nodes_df['shape_text'] = nodes_df['name']
        else:
            nodes_df = pd.DataFrame(columns=['id', 'name', 'type', 'lane', 'process', 'visio_type', 'shape_text'])
        
        # Create DataFrame for edges
        if self.edges:
            edges_df = pd.DataFrame(self.edges)
            edges_df['visio_type'] = edges_df['type'].apply(self.map_edge_type_to_visio)
            # Edges in Visio need source and target shape names
            edges_df['source_shape'] = edges_df['source'].apply(
                lambda src: self.get_node_name_by_id(src, nodes_df))
            edges_df['target_shape'] = edges_df['target'].apply(
                lambda tgt: self.get_node_name_by_id(tgt, nodes_df))
        else:
            edges_df = pd.DataFrame(columns=['id', 'type', 'name', 'source', 'target', 'visio_type', 'source_shape', 'target_shape'])
        
        # Create DataFrame for swimlanes
        if self.swimlanes:
            swimlanes_df = pd.DataFrame(self.swimlanes)
        else:
            swimlanes_df = pd.DataFrame(columns=['id', 'name', 'process'])
        
        return nodes_df, edges_df, swimlanes_df
    
    def get_node_name_by_id(self, node_id, nodes_df):
        """Get node name by ID for edge connections"""
        if nodes_df.empty:
            return ""
        
        matching_nodes = nodes_df[nodes_df['id'] == node_id]
        if not matching_nodes.empty:
            return matching_nodes.iloc[0]['name']
        
        # Handle data objects and other non-standard nodes
        if node_id and node_id.startswith('data_'):
            return node_id.replace('data_', '')
        
        return node_id
    
    def map_node_type_to_visio(self, node_type):
        """Map BPL node types to Visio shape types"""
        type_mapping = {
            'task': 'Process',
            'send': 'Send',
            'receive': 'Receive',
            'gateway': 'Decision',
            'branch': 'Process',
            'event': 'Event',
            'comment': 'Note'
        }
        return type_mapping.get(node_type, 'Process')
    
    def map_edge_type_to_visio(self, edge_type):
        """Map BPL edge types to Visio connector types"""
        type_mapping = {
            'sequenceFlow': 'Sequence',
            'messageFlow': 'Message',
            'dataAssociation': 'Data'
        }
        return type_mapping.get(edge_type, 'Sequence')
    
    def generate_excel(self, output_path):
        """Generate Visio-compatible Excel file"""
        nodes_df, edges_df, swimlanes_df = self.create_visio_dataframes()
        
        # Create a workbook directly with openpyxl instead of using pandas ExcelWriter
        wb = Workbook()
        ws = wb.active
        ws.title = "Process Diagram"
        
        # Write header
        header = ["Business Process Diagram"]
        ws.append(header)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        header_cell = ws.cell(row=1, column=1)
        header_cell.font = Font(size=14, bold=True)
        
        # Add info
        ws.append(["Generated from BPL AST"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        
        # Add empty row
        ws.append([])
        
        # Swimlanes section
        ws.append(["Swimlanes"])
        swimlanes_header = ["ID", "Name", "Process"]
        ws.append(swimlanes_header)
        
        # Add swimlanes data
        row = 6
        for _, lane in swimlanes_df.iterrows():
            ws.append([lane['id'], lane['name'], lane['process']])
            row += 1
        
        # Add empty row
        ws.append([])
        row += 1
        
        # Nodes section
        ws.append(["Shapes"])
        row += 1
        nodes_header = ["ID", "Name", "Type", "Lane", "Visio Shape"]
        ws.append(nodes_header)
        row += 1
        
        # Add nodes data
        for _, node in nodes_df.iterrows():
            ws.append([
                node['id'], 
                node['name'], 
                node['type'],
                node['lane'], 
                node['visio_type']
            ])
            row += 1
        
        # Add empty row
        ws.append([])
        row += 1
        
        # Edges section
        ws.append(["Connections"])
        row += 1
        edges_header = ["ID", "Type", "Name", "Source", "Target", "Visio Type"]
        ws.append(edges_header)
        row += 1
        
        # Add edges data
        for _, edge in edges_df.iterrows():
            ws.append([
                edge['id'],
                edge['type'],
                edge['name'],
                edge['source'],
                edge['target'],
                edge['visio_type']
            ])
            row += 1
        
        # Add styling
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save the workbook
        wb.save(output_path)
        
        print(f"Excel file generated at: {output_path}")
        return output_path

def main():
    if len(sys.argv) < 3:
        print("Usage: python ast_to_visio.py input.bpl-ast.json output.xlsx")
        sys.exit(1)
    
    input_json = sys.argv[1]
    output_xlsx = sys.argv[2]
    
    if not os.path.exists(input_json):
        print(f"Error: Input file {input_json} not found.")
        sys.exit(1)
    
    try:
        with open(input_json, 'r') as f:
            ast_data = json.load(f)
        
        converter = BplAstToVisioConverter(ast_data)
        converter.process_ast()
        converter.generate_excel(output_xlsx)
        
        print(f"Successfully converted {input_json} to Visio format at {output_xlsx}")
    except Exception as e:
        print(f"Error processing the AST: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()